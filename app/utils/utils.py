import datetime
import json
import re
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def deep_get(dct, keys, default=None):
    """Safely get nested dict value by a list or single key."""
    if not isinstance(keys, (list, tuple)):
        keys = [keys]
    current = dct
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key, default)
    return current

def parse_content(content):
    """Parse content to extract JSON data."""
    if not content:
        return None
    try:
        data = json.loads(content)
        return deep_get(data, "data")
    except json.JSONDecodeError:
        regex = r"\"data\":({\"n.*?),\"extensions\""
        match = re.search(regex, content, re.DOTALL)
        if not match:
            return None
        return json.loads(match.group(1))

def clean_graphql_response(response_text: str) -> str:
    """Remove unwanted JSON fragments from GraphQL response."""
    unwanted_pattern = r'\{"label":"VideoPlayerRelay_video\$defer\$InstreamVideoAdBreaksPlayer_video".*\}'
    cleaned_response = re.sub(unwanted_pattern, '', response_text, flags=re.DOTALL)
    return cleaned_response.rstrip('}, \n')

def parse_graphql_comment_replies(response, response_type):
    """Parse GraphQL response for comments or replies, based on response_type ('comments' or 'replies')."""
    try:
        data = json.loads(response)
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON response for {response_type}")
        return {"results": [], "page_info": {"end_cursor": None, "has_next_page": False}}

    # Define paths based on response type
    if response_type == "comments":
        container_paths = [
            ["node", "comment_rendering_instance_for_feed_location", "comments"],
            ["node", "feedback", "comment_rendering_instance_for_feed_location", "comments"]
        ]
        node_fields = {
            "author": ["user", "name"],
            "author_id": ["user", "id"],
            "text": ["body", "text"],
            "comment_id": ["id"],
            "legacy_fbid": ["legacy_fbid"],
            "created_time": ["created_time"],
            "reaction_count": ["reactors", "count_reduced"],
            "reply_count": ["feedback", "replies_fields", "total_count"],
            "feedback_id": ["feedback", "id"],
            "expansion_token": ["feedback", "expansion_info", "expansion_token"]
        }
    elif response_type == "replies":
        container_paths = [["data", "node", "replies_connection"]]
        node_fields = {
            "author": ["author", "name"],
            "author_id": ["author", "id"],
            "text": ["body", "text"],
            "reply_id": ["id"],
            "legacy_fbid": ["legacy_fbid"],
            "created_time": ["created_time"],
            "reaction_count": ["feedback", "reactors", "count_reduced"],
            "profile_picture": ["author", "profile_picture_depth_0", "uri"],
            "gender": ["author", "gender"],
            "feedback_id": ["feedback", "id"]
        }
    else:
        raise ValueError(f"Invalid response_type: {response_type}")

    # Find the container
    container = None
    for path in container_paths:
        container = deep_get(data, path)
        if isinstance(container, dict) and "edges" in container:
            break
    if not (isinstance(container, dict) and "edges" in container):
        print(f"Error: No {response_type}_connection found in response")
        return {"results": [], "page_info": {"end_cursor": None, "has_next_page": False}}

    # Extract pagination info
    page_info = deep_get(container, "page_info", {})
    end_cursor = deep_get(page_info, "end_cursor")
    has_next_page = deep_get(page_info, "has_next_page", False)

    results = []
    for edge in deep_get(container, "edges", []):
        node = deep_get(edge, "node", {})
        result = {}
        
        # Extract fields based on node_fields mapping
        for field, path in node_fields.items():
            value = deep_get(node, path)
            if field == "created_time" and value:
                if isinstance(value, (int, float)):
                    value = datetime.datetime.fromtimestamp(
                        value, tz=datetime.timezone.utc
                    ).strftime('%Y-%m-%d %H:%M:%S UTC')
            elif field == "reply_count" and value is None:
                value = 0
            result[field] = value

        # Add replies list for comments
        if response_type == "comments":
            result["replies"] = []

        results.append(result)

    return {
        "results": results,
        "page_info": {
            "end_cursor": end_cursor,
            "has_next_page": has_next_page
        }
    }

def parse_facebook_post(data, comment_data=None):
    """Parse Facebook post JSON to extract key info."""
    try:
        post_info = {
            "content": None,
            "post_url": None,
            "article_url": None,
            "reactions": {"total_count": 0, "details": []},
            "shares": 0,
            "comments": {"total_count": 0, "details": []},
            "post_id": None,
            "creation_time": None,
            "author": {"name": None, "id": None, "profile_url": None, "profile_picture": None},
            "privacy_scope": None,
            "attachment": {"title": None, "image_url": None, "media_id": None}
        }

        node = deep_get(data, "node", {})
        cs = deep_get(node, "comet_sections", {})

        # Timestamp
        raw_ts = deep_get(cs, ["timestamp", "story", "creation_time"])
        if raw_ts:
            try:
                ts = float(raw_ts)
                post_info["creation_time"] = datetime.datetime.fromtimestamp(
                    ts, tz=datetime.timezone.utc
                ).strftime('%Y-%m-%d %H:%M:%S UTC')
            except Exception:
                post_info["creation_time"] = None

        # Main content
        story = deep_get(cs, ["content", "story"], {})
        post_info["content"] = deep_get(story, ["message", "text"])
        post_info["post_url"] = deep_get(story, "wwwURL")
        post_info["post_id"] = deep_get(node, "id")

        # External article link
        attachments = deep_get(story, "attachments", [])
        if attachments:
            att = attachments[0] or {}
            article = deep_get(att, ["comet_footer_renderer", "target", "external_url"])
            if not article:
                article = deep_get(att, [
                    "styles", "attachment",
                    "story_attachment_link_renderer", "attachment",
                    "web_link", "url"
                ])
            post_info["article_url"] = article

        # Reactions & shares
        ufi = deep_get(cs, [
            "feedback", "story", "story_ufi_container",
            "story", "feedback_context",
            "feedback_target_with_context",
            "comet_ufi_summary_and_actions_renderer"
        ], {})
        fb = deep_get(ufi, "feedback", {})
        post_info["reactions"]["total_count"] = deep_get(fb, ["reaction_count", "count"], 0)
        for edge in deep_get(fb, ["top_reactions", "edges"], []):
            n = deep_get(edge, "node", {})
            post_info["reactions"]["details"].append({
                "reaction": deep_get(n, "localized_name"),
                "count": deep_get(edge, "reaction_count")
            })
        post_info["shares"] = deep_get(fb, ["share_count", "count"], 0)

        # Native comments
        comment_list = deep_get(ufi, [
            "feedback_target_with_context",
            "comment_list_renderer",
            "feedback",
            "comment_rendering_instance_for_feed_location",
            "comments"
        ], {})
        post_info["comments"]["total_count"] = deep_get(comment_list, "total_count", 0)
        for edge in deep_get(comment_list, "edges", []):
            cn = deep_get(edge, "node", {})
            auth = deep_get(cn, "author", {})
            fb_c = deep_get(cn, "feedback", {})
            post_info["comments"]["details"].append({
                "author": deep_get(auth, "name"),
                "author_id": deep_get(auth, "id"),
                "text": deep_get(cn, ["body", "text"]),
                "comment_id": deep_get(cn, "id"),
                "legacy_fbid": deep_get(cn, "legacy_fbid"),
                "created_time": deep_get(cn, "created_time"),
                "reaction_count": deep_get(fb_c, ["reactors", "count_reduced"]),
                "reply_count": deep_get(fb_c, ["replies_fields", "total_count"], 0),
                "replies": [],
                "feedback_id": deep_get(fb_c, "id")
            })

        # Additional comments from comment_data
        if comment_data:
            for comment_set in comment_data:
                post_info["comments"]["total_count"] = max(
                    post_info["comments"]["total_count"], len(comment_set)
                )
                for comment in comment_set:
                    cid = deep_get(comment, "comment_id")
                    if not any(c["comment_id"] == cid for c in post_info["comments"]["details"]):
                        post_info["comments"]["details"].append({
                            "author": deep_get(comment, "author"),
                            "author_id": deep_get(comment, "author_id"),
                            "text": deep_get(comment, "text"),
                            "comment_id": cid,
                            "legacy_fbid": deep_get(comment, "legacy_fbid"),
                            "created_time": deep_get(comment, "created_time"),
                            "reaction_count": deep_get(comment, "reaction_count"),
                            "reply_count": deep_get(comment, "reply_count", 0),
                            "replies": deep_get(comment, "replies", []),
                            "feedback_id": deep_get(comment, "feedback_id")
                        })

        # Author info & privacy
        ctx_md = deep_get(cs, ["context_layout", "story", "comet_sections", "metadata"], [])
        if ctx_md:
            ts_meta = deep_get(ctx_md[0], ["story", "creation_time"])
            if ts_meta:
                post_info["creation_time"] = ts_meta
        actors = deep_get(story, "actors", [])
        if actors:
            a = actors[0] or {}
            post_info["author"].update({
                "name": deep_get(a, "name"),
                "id": deep_get(a, "id"),
                "profile_url": deep_get(a, "url"),
                "profile_picture": deep_get(a, ["profile_picture", "uri"])
            })
        post_info["privacy_scope"] = deep_get(
            deep_get(ctx_md[0] if ctx_md else {}, "story", {}),
            ["privacy_scope", "description"]
        )

        # Final attachment details
        if attachments:
            att = attachments[0] or {}
            styles = deep_get(att, "styles", {})
            media = deep_get(styles, ["attachment", "media"], {})
            post_info["attachment"].update({
                "title": deep_get(styles, ["title_with_entities", "text"]),
                "image_url": deep_get(media, ["large_share_image", "uri"]),
                "media_id": deep_get(media, "id")
            })

        return post_info
    except Exception as e:
        print(f"parse_facebook_post failed: {e}")
        return None

def extract_post_id_from_html(html_content):
    """Extract the post ID from HTML content."""
    match = re.search(r'"post_id":"([0-9]{10,})"', html_content)
    if match:
        return match.group(1)
    match = re.search(r'"feedback_id":"ZmVlZGJhY2s6([0-9]{10,})"', html_content)
    if match:
        return match.group(1)
    raise ValueError("Could not extract post ID from HTML")

def encode_feedback_id(post_id):
    """Encode the feedback ID in Base64 for GraphQL query."""
    feedback_str = f"feedback:{post_id}"
    return base64.b64encode(feedback_str.encode()).decode()

async def parse_html_for_params(html_content, required_params):
    """Extract specific GraphQL parameters from HTML content."""
    params = {}
    if not required_params:
        return params

    script_pattern = re.compile(r'<script type="application/json"[^>]*>(.*?)</script>', re.DOTALL)
    scripts = script_pattern.findall(html_content)

    for script_content in scripts:
        try:
            data = json.loads(script_content)
            def extract_params(obj):
                if isinstance(obj, dict):
                    for key, value in obj.items():
                        if key in required_params:
                            params[key] = str(value)
                        elif key == "consistency" and isinstance(value, dict) and "rev" in value and "__rev" in required_params:
                            params["__rev"] = str(value["rev"])
                        if isinstance(value, (dict, list)):
                            extract_params(value)
                elif isinstance(obj, list):
                    for item in obj:
                        extract_params(item)
            extract_params(data)
            if all(param in params for param in required_params):
                break
        except json.JSONDecodeError:
            continue
    return params

def save_to_excel(post_data, filename):
    """Save post and comments data to an Excel file with two sheets."""
    # Prepare data for Post sheet
    post_info = {
        "post_id": post_data.get("post_id"),
        "content": post_data.get("content"),
        "post_url": post_data.get("post_url"),
        "article_url": post_data.get("article_url"),
        "creation_time": post_data.get("creation_time"),
        "author_name": post_data.get("author", {}).get("name"),
        "author_id": post_data.get("author", {}).get("id"),
        "author_profile_url": post_data.get("author", {}).get("profile_url"),
        "author_profile_picture": post_data.get("author", {}).get("profile_picture"),
        "privacy_scope": post_data.get("privacy_scope"),
        "reactions_total_count": post_data.get("reactions", {}).get("total_count"),
        "shares": post_data.get("shares"),
        "comments_total_count": post_data.get("comments", {}).get("total_count"),
        "attachment_title": post_data.get("attachment", {}).get("title"),
        "attachment_image_url": post_data.get("attachment", {}).get("image_url"),
        "attachment_media_id": post_data.get("attachment", {}).get("media_id")
    }
    post_df = pd.DataFrame([post_info])

    # Prepare data for Comments sheet
    comments = post_data.get("comments", {}).get("details", [])
    flat_comments = []
    for comment in comments:
        flat_comments.append({
            "author": comment.get("author"),
            "author_id": comment.get("author_id"),
            "text": comment.get("text"),
            "comment_id": comment.get("comment_id"),
            "legacy_fbid": comment.get("legacy_fbid"),
            "created_time": comment.get("created_time"),
            "reaction_count": comment.get("reaction_count"),
            "reply_count": comment.get("reply_count"),
            "feedback_id": comment.get("feedback_id")
        })
        for reply in comment.get("replies", []):
            flat_comments.append({
                "author": reply.get("author"),
                "author_id": reply.get("author_id"),
                "text": reply.get("text"),
                "comment_id": reply.get("reply_id"),
                "legacy_fbid": reply.get("legacy_fbid"),
                "created_time": reply.get("created_time"),
                "reaction_count": reply.get("reaction_count"),
                "reply_count": 0,
                "feedback_id": reply.get("feedback_id")
            })
    comments_df = pd.DataFrame(flat_comments if flat_comments else [{"author": "No comments found", "comment_id": None}])

    # Create Excel file
    workbook = Workbook()

    # Create Post sheet
    ws_post = workbook.create_sheet(title="Post")
    for row in dataframe_to_rows(post_df, index=False, header=True):
        ws_post.append(row)

    # Create Comments sheet
    ws_comments = workbook.create_sheet(title="Comments")
    for row in dataframe_to_rows(comments_df, index=False, header=True):
        ws_comments.append(row)

    # Remove default sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Save to file
    workbook.save(filename)